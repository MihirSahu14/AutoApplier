"""Export DB to a multi-sheet Excel workbook for manual tracking."""
import json
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from . import db

OUT = Path(__file__).resolve().parent.parent / "data" / "JobTracker.xlsx"

STATUS_OPTIONS = [
    "Not started", "To apply", "Applied", "Phone screen", "Technical",
    "Onsite", "Offer", "Rejected", "Withdrawn", "Ghosted",
]
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _style_header(ws, n_cols: int):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def _autosize(ws, max_w: int = 60):
    for col_idx, col in enumerate(ws.columns, start=1):
        longest = 0
        for cell in col:
            if cell.value is None:
                continue
            longest = max(longest, min(len(str(cell.value)), max_w))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, longest + 2)


def _read_existing_status() -> dict[int, dict]:
    """Preserve user-edited status/notes across re-exports."""
    if not OUT.exists():
        return {}
    try:
        wb = load_workbook(OUT, read_only=True)
    except Exception:
        return {}
    if "Applications" not in wb.sheetnames:
        return {}
    ws = wb["Applications"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = dict(zip(headers, row))
        jid = rec.get("Job ID")
        if isinstance(jid, int):
            out[jid] = {
                "Status": rec.get("Status") or "Not started",
                "Applied On": rec.get("Applied On"),
                "Resume Path": rec.get("Resume Path"),
                "Cover Letter Path": rec.get("Cover Letter Path"),
                "Contact Name": rec.get("Contact Name"),
                "Contact Email": rec.get("Contact Email"),
                "Email Sent": rec.get("Email Sent"),
                "Notes": rec.get("Notes"),
            }
    return out


def export() -> Path:
    OUT.parent.mkdir(parents=True, exist_ok=True)
    preserved = _read_existing_status()

    with db.connect() as c:
        rows = c.execute(
            "SELECT j.id, j.source, j.company, j.title, j.location, j.url, "
            "j.posted_at, j.ingested_at, j.description, "
            "s.score, s.fit_summary, s.disqualified, s.disqualify_reason, "
            "s.matched_skills, s.missing_skills "
            "FROM jobs j LEFT JOIN scores s ON s.job_id = j.id "
            "ORDER BY COALESCE(s.score, -1) DESC, j.id DESC"
        ).fetchall()

    wb = Workbook()

    # --- Applications: the active tracker, only qualified jobs >= 60 ---
    ws_app = wb.active
    ws_app.title = "Applications"
    app_headers = [
        "Job ID", "Score", "Company", "Title", "Location", "Source", "URL",
        "Status", "Applied On", "Resume Path", "Cover Letter Path",
        "Contact Name", "Contact Email", "Email Sent", "Notes",
        "Fit Summary", "Matched Skills", "Missing Skills",
    ]
    ws_app.append(app_headers)
    _style_header(ws_app, len(app_headers))

    for r in rows:
        if r["score"] is None or r["disqualified"] or (r["score"] or 0) < 60:
            continue
        prev = preserved.get(r["id"], {})
        ws_app.append([
            r["id"], r["score"], r["company"], r["title"], r["location"],
            r["source"], r["url"],
            prev.get("Status") or "Not started",
            prev.get("Applied On"),
            prev.get("Resume Path"),
            prev.get("Cover Letter Path"),
            prev.get("Contact Name"),
            prev.get("Contact Email"),
            prev.get("Email Sent"),
            prev.get("Notes"),
            r["fit_summary"],
            ", ".join(json.loads(r["matched_skills"] or "[]")),
            ", ".join(json.loads(r["missing_skills"] or "[]")),
        ])

    # data validation dropdown for Status
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(STATUS_OPTIONS) + '"',
        allow_blank=True,
    )
    dv.add(f"H2:H{ws_app.max_row}")
    ws_app.add_data_validation(dv)
    _autosize(ws_app)

    # --- All Jobs: full ledger including unscored / low score ---
    ws_all = wb.create_sheet("All Jobs")
    all_headers = [
        "Job ID", "Score", "Disqualified", "Reason", "Company", "Title",
        "Location", "Source", "URL", "Posted", "Ingested", "Fit Summary",
    ]
    ws_all.append(all_headers)
    _style_header(ws_all, len(all_headers))
    for r in rows:
        ws_all.append([
            r["id"], r["score"],
            "Yes" if r["disqualified"] else ("No" if r["score"] is not None else ""),
            r["disqualify_reason"], r["company"], r["title"], r["location"],
            r["source"], r["url"], r["posted_at"], r["ingested_at"],
            r["fit_summary"],
        ])
    _autosize(ws_all)

    # --- Disqualified: hard-filter rejects, for review ---
    ws_dq = wb.create_sheet("Disqualified")
    dq_headers = ["Job ID", "Score", "Company", "Title", "Reason", "URL"]
    ws_dq.append(dq_headers)
    _style_header(ws_dq, len(dq_headers))
    for r in rows:
        if r["disqualified"]:
            ws_dq.append([r["id"], r["score"], r["company"], r["title"],
                          r["disqualify_reason"], r["url"]])
    _autosize(ws_dq)

    # --- Summary sheet ---
    ws_sum = wb.create_sheet("Summary", 0)  # first tab
    ws_sum.append(["Metric", "Value"])
    _style_header(ws_sum, 2)
    total = len(rows)
    scored = sum(1 for r in rows if r["score"] is not None)
    dq = sum(1 for r in rows if r["disqualified"])
    qualified = sum(1 for r in rows if r["score"] is not None and not r["disqualified"])
    high = sum(1 for r in rows if (r["score"] or 0) >= 80 and not r["disqualified"])
    metrics = [
        ("Total jobs ingested", total),
        ("Scored", scored),
        ("Qualified (passed hard filters)", qualified),
        ("Disqualified", dq),
        ("Strong fit (>=80)", high),
        ("In Applications tab (>=60)", ws_app.max_row - 1),
    ]
    for k, v in metrics:
        ws_sum.append([k, v])
    _autosize(ws_sum)

    wb.save(OUT)
    return OUT
